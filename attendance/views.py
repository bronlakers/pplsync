import calendar
import datetime as dt
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from core.utils import group_required, in_group
from core.models import Employee
from .forms import AttendanceUploadForm
from .models import AttendanceImport, AttendanceRecord
from .utils import minutes_between, clamp_nonneg

EXPECTED_COLUMNS = ["employee_code", "date", "check_in", "check_out", "status"]

def _parse_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        return dt.datetime.strptime(v.strip(), "%Y-%m-%d").date()
    raise ValueError("Invalid date")

def _parse_time(v):
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.time()
    if isinstance(v, dt.time):
        return v
    if isinstance(v, str):
        return dt.datetime.strptime(v.strip(), "%H:%M").time()
    raise ValueError("Invalid time")

@login_required
@group_required("HR")
def upload(request):
    form = AttendanceUploadForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        imp = AttendanceImport.objects.create(
            uploaded_by=request.user,
            file=form.cleaned_data["file"],
            status="processing",
        )
        errors = []
        created = 0
        updated = 0
        try:
            wb = load_workbook(imp.file.path)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            header_map = {h: i for i, h in enumerate(headers)}

            missing = [c for c in EXPECTED_COLUMNS if c not in header_map]
            if missing:
                raise ValueError(f"Missing columns: {', '.join(missing)}")

            with transaction.atomic():
                for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    try:
                        employee_code = str(row[header_map["employee_code"]].value).strip()
                        date = _parse_date(row[header_map["date"]].value)
                        check_in = _parse_time(row[header_map["check_in"]].value)
                        check_out = _parse_time(row[header_map["check_out"]].value)
                        status = (str(row[header_map["status"]].value).strip().lower() or "present")

                        employee = Employee.objects.get(employee_code=employee_code)
                        shift = employee.default_shift

                        late = early = worked = 0
                        if status == "present" and shift and check_in and check_out:
                            late = clamp_nonneg(minutes_between(shift.start_time, check_in) - shift.grace_minutes_late)
                            early = clamp_nonneg(minutes_between(check_out, shift.end_time) - shift.grace_minutes_early)
                            worked = max(0, minutes_between(check_in, check_out))

                        obj, created_flag = AttendanceRecord.objects.update_or_create(
                            employee=employee,
                            date=date,
                            defaults=dict(
                                check_in=check_in,
                                check_out=check_out,
                                status=status,
                                shift=shift,
                                late_minutes=late,
                                early_minutes=early,
                                worked_minutes=worked,
                                source_import=imp,
                            )
                        )
                        created += 1 if created_flag else 0
                        updated += 0 if created_flag else 1

                    except Exception as e:
                        errors.append(f"Row {r_idx}: {e}")

            imp.status = "failed" if errors else "success"
            imp.error_report = "\n".join(errors)
            imp.save()

            if errors:
                messages.warning(request, f"Imported with {len(errors)} errors. Open import record in admin to view errors.")
            else:
                messages.success(request, f"Imported successfully. Created {created}, updated {updated}.")
            return redirect("attendance_month_view")

        except Exception as e:
            imp.status = "failed"
            imp.error_report = str(e)
            imp.save()
            messages.error(request, f"Import failed: {e}")
            return redirect("attendance_upload")

    return render(request, "attendance/upload.html", {"form": form})

@login_required
def month_view(request):
    if not (request.user.is_superuser or in_group(request.user, "HR")):
        return redirect("attendance_my_history")

    today = dt.date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    first = dt.date(year, month, 1)
    last = dt.date(year, month, calendar.monthrange(year, month)[1])

    qs = AttendanceRecord.objects.select_related("employee").filter(date__range=(first, last)).order_by("date", "employee__full_name")
    exceptions = qs.filter(late_minutes__gt=0) | qs.filter(early_minutes__gt=0) | qs.filter(status__in=["leave", "wfh", "absent", "halfday"])

    ctx = {
        "year": year,
        "month": month,
        "month_name": calendar.month_name[month],
        "records": qs[:500],
        "exceptions": exceptions[:500],
        "prev_year": (first - dt.timedelta(days=1)).year,
        "prev_month": (first - dt.timedelta(days=1)).month,
        "next_year": (last + dt.timedelta(days=1)).year,
        "next_month": (last + dt.timedelta(days=1)).month,
        "is_hr_view": True,
    }
    return render(request, "attendance/month_view.html", ctx)

@login_required
def my_history(request):
    employee = Employee.objects.filter(user=request.user).first()
    if not employee:
        messages.error(request, "Your user is not linked to an Employee record. Ask admin to link it in Admin > Employees.")
        return redirect("dashboard")

    today = dt.date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    first = dt.date(year, month, 1)
    last = dt.date(year, month, calendar.monthrange(year, month)[1])

    qs = AttendanceRecord.objects.filter(employee=employee, date__range=(first, last)).order_by("date")
    ctx = {
        "employee": employee,
        "year": year,
        "month": month,
        "month_name": calendar.month_name[month],
        "records": qs,
        "prev_year": (first - dt.timedelta(days=1)).year,
        "prev_month": (first - dt.timedelta(days=1)).month,
        "next_year": (last + dt.timedelta(days=1)).year,
        "next_month": (last + dt.timedelta(days=1)).month,
        "is_hr_view": False,
    }
    return render(request, "attendance/my_history.html", ctx)
